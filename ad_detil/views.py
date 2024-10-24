from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect

from openpyxl.workbook import Workbook


from ad_detil.models import Tbl_cat, Sub_category
from advocate.models import Tbl_adv
from customer.models import Tbl_cust, Payment_Confirmation, booking
from staff.models import Tbl_staff

def demologout(request):
    logout(request)
    return redirect('/')
def admindash(requset):

    if requset.user.is_advocate:
        staff=0
        adv = len(booking.objects.filter(adv__u=requset.user))
        cust = len(booking.objects.filter(adv__u=requset.user))
        staff = len(booking.objects.filter(adv__u=requset.user))
        book = 0

    else:
        adv = len(Tbl_adv.objects.all())
        cust= len(Tbl_cust.objects.all())
        staff = len(Tbl_staff.objects.all())
        book = len(booking.objects.all())


    return render(requset,'dash/index.html',{'adv':adv,'cust':cust,'staff':staff,'book':book})

def staffreg(requset):
    return render(requset,'staffreg.html')


def category(requset):
    if requset.method == 'POST':
        catname=requset.POST['cname']
        cdes = requset.POST['cdes']
        cat=Tbl_cat(Cat_name=catname,Cat_desc=cdes)
        cat.save()
        return redirect('catetbl')
        messages.info(requset, "Add success")
    return render(requset,'dash/democat.html')

def subcat(requset):
    if requset.method == 'POST':
        subname = requset.POST['subname']
        subdes = requset.POST['subdes']
        category=requset.POST['cat']
        c=Tbl_cat.objects.get(id=category)
        subcat = Sub_category(sub_name=subname, des=subdes,category=c)
        subcat.save()
        messages.info(requset, "Add success")
        return redirect('subtbl')

    cat=Tbl_cat.objects.all()
    return render(requset,'dash/demosubcat.html',{'c':cat})


@login_required(login_url='login')
def catetbl(requset):
    cat = Tbl_cat.objects.all()

    return render(requset, 'dash/catetbl.html',{'category':cat})

@login_required(login_url='login')
def subtbl(requset):
    sub = Sub_category.objects.all()

    return render(requset, 'dash/subtbl.html',{'subcat':sub})

@login_required(login_url='login')
def stafftbl(requset):
    staff = Tbl_staff.objects.all()

    return render(requset, 'dash/stafftbl.html',{'staff':staff})


from django.template.loader import get_template
from django.conf import settings
from xhtml2pdf import pisa
from io import BytesIO

def staffpdf(request):
    # Get data for your template, you can replace it with your own dynamic data retrieval logic

    staff = Tbl_staff.objects.filter(S_status=True)
    context = {
        'staff': staff
    }

    # Render the template
    template = get_template('staff_pdf_template.html')
    html_content = template.render(context)

    # Create a PDF file
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'filename="your_pdf_filename.pdf"'
        return response

    return HttpResponse('Error generating PDF: %s' % pdf.err)



@login_required(login_url='login')
def advtbl(request):
    if request.method == 'POST':
        start_date = request.POST['sdate']
        end_date = request.POST['edate']
        adv = Tbl_adv.objects.filter(Adv_join_date__gte=start_date, Adv_join_date__lte=end_date)
    else:
        adv = Tbl_adv.objects.all()

    return render(request, 'dash/advtbl.html',{'adv':adv})

@login_required(login_url='login')
def custtbl(requset):
    cust = Tbl_cust.objects.all()
    return render(requset, 'dash/custtbl.html', {'cust': cust})

@login_required(login_url='login')
def bk(requset):

    return render(requset, 'dash/custtbl.html', {'cust': cust})

@login_required(login_url='login')
def paytbl(request):
    user = request.user
    t = Tbl_adv.objects.get(u=user)
    pay= Payment_Confirmation.objects.filter(book__adv=t)
    return render(request,'dash/paytbl.html',{'pay':pay})

@login_required(login_url='login')
def booktbl(request):
    user=request.user
    t=Tbl_adv.objects.get(u=user)
    book= booking.objects.filter(adv=t)
    return render(request,'dash/booktbl.html',{'book':book})



@login_required(login_url='login')
def edit(request,pid):
    cat = Tbl_cat.objects.get(id=pid)
    if request.method =='POST':
        cname=request.POST['cname']
        cdes= request.POST['cdes']
        cat.Cat_name=cname
        cat.Cat_desc=cdes
        cat.save()
        return redirect('catetbl')
    else:
        cat = Tbl_cat.objects.get(id=pid)

    return render(request,'dash/cedit.html',{'cat':cat})


def activate(request,pid):
    staff = Tbl_staff.objects.get(id=pid)
    staff.S_status=True
    staff.save()

    return redirect('stafftbl')


def deactivate(request, pid):
    staff = Tbl_staff.objects.get(id=pid)
    staff.S_status = False
    staff.save()

    return redirect('stafftbl')

def advactivate(request,pid):
    adv = Tbl_adv.objects.get(id=pid)
    adv.Adv_status=True
    adv.save()

    return redirect('advtbl')


def advdeactivate(request, pid):
    adv = Tbl_adv.objects.get(id=pid)
    adv.Adv_status = False
    adv.save()

    return redirect('advtbl')



def custactivate(request,pid):
    cust = Tbl_cust.objects.get(id=pid)
    cust.C_status=True
    cust.save()

    return redirect('custtbl')


def custdeactivate(request, pid):
    cust = Tbl_cust.objects.get(id=pid)
    cust.C_status = False
    cust.save()

    return redirect('custtbl')
@login_required(login_url='login')
def subedit(request,pid):
    cat = Sub_category.objects.get(id=pid)
    if request.method =='POST':
        cname=request.POST['subname']
        cdes= request.POST['subdes']
        cs = request.POST['scat']
        cat.sub_name=cname
        cat.des=cdes
        g=Tbl_cat.objects.get(id=cs)
        cat.category=g

        cat.save()
        return redirect('subtbl')
    else:
        cat = Sub_category.objects.get(id=pid)
    t=Tbl_cat.objects.all()
    data_from_database = Tbl_cat.objects.first()

    return render(request,'dash/subedit.html',{'cat':cat,'c':t,'selected_category_name':data_from_database})


def database_to_excel__staff(request):
    # Fetch data from the Customer model
    queryset = Tbl_staff.objects.all()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = ['S_fname', 'S_lname', 'S_gen', 'S_email', 'S_ph']
    ws.append(headers)

    # Write data to the worksheet
    for customer in queryset:
        data = []
        for header in headers:
            # Handle the ForeignKey field (user) by extracting the username
            if header == 'user':
                user_value = getattr(customer, header)
                if user_value:
                    data.append(user_value.username)
                else:
                    data.append(None)
            else:
                data.append(getattr(customer, header))

        ws.append(data)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Excel.xlsx'
    wb.save(response)

    return response

def database_to_excel_customer(request):
    # Fetch data from the Customer model
    queryset = Tbl_cust.objects.all()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = ['C_fname', 'C_lname', 'C_gen', 'C_email', 'C_ph']
    ws.append(headers)

    # Write data to the worksheet
    for customer in queryset:
        data = []
        for header in headers:
            # Handle the ForeignKey field (user) by extracting the username
            if header == 'user':
                user_value = getattr(customer, header)
                if user_value:
                    data.append(user_value.username)
                else:
                    data.append(None)
            else:
                data.append(getattr(customer, header))

        ws.append(data)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Excel.xlsx'
    wb.save(response)

    return response


def database_to_excel_advocate(request):
    # Fetch data from the Customer model
    queryset = Tbl_adv.objects.all()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = ['Adv_fname', 'Adv_lname', 'Adv_gen', 'Adv_email', 'Adv_ph']
    ws.append(headers)

    # Write data to the worksheet
    for customer in queryset:
        data = []
        for header in headers:
            # Handle the ForeignKey field (user) by extracting the username
            if header == 'user':
                user_value = getattr(customer, header)
                if user_value:
                    data.append(user_value.username)
                else:
                    data.append(None)
            else:
                data.append(getattr(customer, header))

        ws.append(data)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Excel.xlsx'
    wb.save(response)

    return response

def database_to_excel_booking(request):
    # Fetch data from the Customer model
    q = booking.objects.all()
    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = ['Customer','Advocate', 'Amount', 'Date']
    ws.append(headers)

    # # Write data to the worksheet
    for i in q:
        data = []
        for header in headers:
            # Handle the ForeignKey field (user) by extracting the username
            if header == 'Customer':
                if getattr(i,'payment') != None:
                    user_value = getattr(i,'customer')
                    if user_value:
                        data.append(user_value.C_fname)
                    else:
                        data.append(None)
                else:
                    data.append(None)
            elif header == 'Advocate':
                if getattr(i, 'payment') != None:
                    data.append(i.adv.Adv_fname)
                else:
                    data.append(None)
            elif header == 'Amount':
                data.append(getattr(i, 'payment'))

            else:
                if getattr(i, 'payment') != None:
                    data.append(str(i.date))
                else:
                    data.append(None)
        ws.append(data)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Excel.xlsx'
    wb.save(response)

    return response



# Advocate view

def adv_view(request):
    adv = Tbl_adv.objects.all()
    return render(request,'dash/adv_aproval.html',{'adv':adv})

# adv deapprove

def adv_deapproval(request,id):
    print('function called')
    adv = Tbl_adv.objects.get(id=id)
    adv.Adv_approval = False
    adv.save()
    return redirect('adview')

# adv approve

def adv_approval(request,id):
    adv = Tbl_adv.objects.get(id=id)
    adv.Adv_approval = True
    adv.save()
    return redirect('adview')


@login_required(login_url='login')
def adpaytbl(request):
    if request.method == 'POST':
        start_date = request.POST['sdate']
        end_date = request.POST['edate']
        pay = Payment_Confirmation.objects.filter(date__gte=start_date,date__lte=end_date)
    else:
        pay=Payment_Confirmation.objects.all()
    return render(request,'dash/adpay.html',{'pay':pay})

@login_required(login_url='login')
def adbooktbl(request):
    if request.method == 'POST':
        start_date = request.POST['sdate']
        end_date = request.POST['edate']
        book = booking.objects.filter(date__gte=start_date, date__lte=end_date)
    else:
        book= booking.objects.all()
    return render(request,'dash/adbook.html',{'book':book})